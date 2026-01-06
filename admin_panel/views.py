from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Count, Prefetch
from django.utils.translation import gettext as _
import markdown
from accounts.models import CustomUser
from children.models import Child
from voting.models import DateGroup, DateOption, TimeSlot, Vote
from .forms import DateGroupForm, DateOptionFormSet, WelcomePageForm
from .models import WelcomePage


def is_admin(user):
    """Check if user is an admin or superuser"""
    return user.is_authenticated and (user.is_admin or user.is_superuser or user.is_staff)


@login_required
@user_passes_test(is_admin)
def dashboard(request):
    """Admin dashboard"""
    date_groups = DateGroup.objects.all().annotate(
        total_votes=Count('date_options__time_slots__votes')
    ).order_by('-created_at')
    
    context = {
        'date_groups': date_groups,
    }
    return render(request, 'admin_panel/dashboard.html', context)


@login_required
@user_passes_test(is_admin)
def date_group_create(request):
    """Create a new date group"""
    if request.method == 'POST':
        form = DateGroupForm(request.POST)
        formset = DateOptionFormSet(request.POST)
        
        if form.is_valid() and formset.is_valid():
            date_group = form.save(commit=False)
            date_group.created_by = request.user
            date_group.save()
            formset.instance = date_group
            formset.save()
            messages.success(request, _('Le groupe de dates "%(title)s" a été créé avec succès !') % {'title': date_group.title})
            return redirect('admin_panel:dashboard')
    else:
        form = DateGroupForm()
        formset = DateOptionFormSet()
    
    context = {
        'form': form,
        'formset': formset,
        'title': _('Créer un groupe de dates'),
    }
    return render(request, 'admin_panel/date_group_form.html', context)


@login_required
@user_passes_test(is_admin)
def date_group_edit(request, pk):
    """Edit an existing date group"""
    date_group = get_object_or_404(DateGroup, pk=pk)
    
    if request.method == 'POST':
        form = DateGroupForm(request.POST, instance=date_group)
        formset = DateOptionFormSet(request.POST, instance=date_group)
        
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            messages.success(request, _('Le groupe de dates "%(title)s" a été mis à jour avec succès !') % {'title': date_group.title})
            return redirect('admin_panel:dashboard')
    else:
        form = DateGroupForm(instance=date_group)
        formset = DateOptionFormSet(instance=date_group)
    
    context = {
        'form': form,
        'formset': formset,
        'title': _('Modifier un groupe de dates'),
        'date_group': date_group,
    }
    return render(request, 'admin_panel/date_group_form.html', context)


@login_required
@user_passes_test(is_admin)
def date_group_delete(request, pk):
    """Delete a date group"""
    date_group = get_object_or_404(DateGroup, pk=pk)
    
    if request.method == 'POST':
        title = date_group.title
        date_group.delete()
        messages.success(request, _('Le groupe de dates "%(title)s" a été supprimé avec succès !') % {'title': title})
        return redirect('admin_panel:dashboard')
    
    context = {
        'date_group': date_group,
    }
    return render(request, 'admin_panel/date_group_confirm_delete.html', context)


@login_required
@user_passes_test(is_admin)
def results_view(request, pk):
    """View detailed voting results for a date group"""
    date_group = get_object_or_404(DateGroup, pk=pk)
    statistics = date_group.get_vote_statistics()
    
    context = {
        'date_group': date_group,
        'statistics': statistics,
    }
    return render(request, 'admin_panel/results.html', context)


@login_required
@user_passes_test(is_admin)
def export_excel(request, pk):
    """Export voting results to Excel - one tab per date, one line per child with yes votes"""
    try:        
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from children.models import Child
        from voting.models import Vote, TimeSlot
        
        thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

        date_group = get_object_or_404(DateGroup, pk=pk)
        
        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        ws = wb.create_sheet(title="Résumé")
        ws.append([date_group.title])
        header1 = ['', 'Matin', '', 'Repas', '', 'Après-midi', '']
        header2 = ['Date', 'Oui', 'Non', 'Oui', 'Non', 'Oui', 'Non']
        ws.append(header1)
        ws.append(header2)
        
        # Get all date options for this group
        date_options = date_group.date_options.all().order_by('date')

        for date_option in date_options:
            date_str = date_option.date.strftime('%d-%b-%Y')
            ws.append([date_str,
                       date_option.time_slots.filter(period='morning', votes__choice='yes').count(),
                       date_option.time_slots.filter(period='morning', votes__choice='no').count(),
                       date_option.time_slots.filter(period='lunch', votes__choice='yes').count(),
                       date_option.time_slots.filter(period='lunch', votes__choice='no').count(),
                       date_option.time_slots.filter(period='afternoon', votes__choice='yes').count(),
                       date_option.time_slots.filter(period='afternoon', votes__choice='no').count()
            ])
        
        ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=3)
        ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=5)
        ws.merge_cells(start_row=2, start_column=6, end_row=2, end_column=7)

        ws.column_dimensions['A'].width = 13

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
        
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        for row in ws.iter_rows(min_row=2, max_row=3, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = Font(bold=True)
        
        for date_option in date_options:
            # Create a sheet for each date
            date_str = date_option.date.strftime('%d-%b-%Y')
            ws = wb.create_sheet(title=date_str)
            
            # Headers
            headers1 = ['', date_str, 'Réservation', '','','arrivée', '', 'départ']
            headers2 = ['#', 'Nom', 'M', 'R', 'AM', 'heure', 'signature', 'heure', 'signature']
            ws.append(headers1)
            ws.append(headers2)

            # Style header
            for cell in ws[1]:
                cell.font = Font(bold=True)

            for cell in ws[2]:
                cell.font = Font(bold=True)
            
            # Get all children who have at least one "yes" vote for this date
            time_slots = date_option.time_slots.all()
            children_with_yes = set()
            
            for time_slot in time_slots:
                yes_votes = Vote.objects.filter(
                    time_slot=time_slot,
                    choice='yes'
                ).select_related('child')
                for vote in yes_votes:
                    children_with_yes.add(vote.child)
            
            # Sort children by age (youngest first)
            # Convert to list and sort by age (ascending = youngest first)
            children_list = sorted(list(children_with_yes), key=lambda c: (c.birth_date.year, c.birth_date.month, c.birth_date.day), reverse=True)
            
            # Get votes for each time slot
            morning_slot = time_slots.filter(period='morning').first()
            lunch_slot = time_slots.filter(period='lunch').first()
            afternoon_slot = time_slots.filter(period='afternoon').first()
            
            # Create vote lookup dictionaries
            morning_votes_y = {}
            lunch_votes_y = {}
            afternoon_votes_y = {}

            if morning_slot:
                for vote in Vote.objects.filter(time_slot=morning_slot, choice='yes').select_related('child'):
                    morning_votes_y[vote.child.id] = True
            
            if lunch_slot:
                for vote in Vote.objects.filter(time_slot=lunch_slot, choice='yes').select_related('child'):
                    lunch_votes_y[vote.child.id] = True
            
            if afternoon_slot:
                for vote in Vote.objects.filter(time_slot=afternoon_slot, choice='yes').select_related('child'):
                    afternoon_votes_y[vote.child.id] = True
            
            # Add rows for each child
            separator = False
            offset = 0
            for i, child in enumerate(children_list):
                if not separator and child.age() > 5:              
                    separator = True
                    offset = i
                    ws.append(['']*9)
                    for cell in ws[ws.max_row]:
                        cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type = "solid")

                child_name_with_age = f"{str(child)} ({child.age()} ans)"
                morning_vote = '✓' if child.id in morning_votes_y else ''
                lunch_vote = '✓' if child.id in lunch_votes_y else ''
                afternoon_vote = '✓' if child.id in afternoon_votes_y else ''
                ws.append([
                    i+1 - offset,
                    child_name_with_age,
                    morning_vote,
                    lunch_vote,
                    afternoon_vote,
                ])

            # Add a "Total" row under the last child row
            total_row = ["Total", ""]
            total_row.append(morning_slot.votes.filter(choice='yes').count())
            total_row.append(lunch_slot.votes.filter(choice='yes').count())
            total_row.append(afternoon_slot.votes.filter(choice='yes').count())
            ws.append(total_row)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            ws.column_dimensions['C'].width = 3.5
            ws.column_dimensions['D'].width = 3.5
            ws.column_dimensions['E'].width = 3.5
            ws.column_dimensions['F'].width = 10
            ws.column_dimensions['G'].width = 27
            ws.column_dimensions['H'].width = 10
            ws.column_dimensions['I'].width = 27
            
            ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=5)
            ws.merge_cells(start_row=1, start_column=6, end_row=1, end_column=7)
            ws.merge_cells(start_row=1, start_column=8, end_row=1, end_column=9)

            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{date_group.title}_results.xlsx"'
        wb.save(response)
        return response
    except ImportError:
        messages.error(request, _('L\'export Excel nécessite openpyxl. Veuillez l\'installer.'))
        return redirect('admin_panel:results', pk=pk)


@login_required
@user_passes_test(is_admin)
def parents_list(request):
    """List all registered parents with their email and children"""
    parents = CustomUser.objects.prefetch_related(
        Prefetch('children', queryset=Child.objects.order_by('last_name', 'first_name'))
    ).filter(is_parent=True).order_by('last_name', 'first_name')
    
    context = {
        'parents': parents,
    }
    return render(request, 'admin_panel/parents_list.html', context)


@login_required
@user_passes_test(is_admin)
def reset_parent_password(request, parent_id):
    """Reset a parent's password to 0000"""
    parent = get_object_or_404(CustomUser, pk=parent_id, is_parent=True)
    
    if request.method == 'POST':
        parent.set_password('0000')
        parent.save()
        messages.success(request, _('Le mot de passe de %(name)s a été réinitialisé à 0000.') % {'name': f"{parent.first_name} {parent.last_name}"})
        return redirect('admin_panel:parents_list')
    
    context = {
        'parent': parent,
    }
    return render(request, 'admin_panel/reset_password_confirm.html', context)


@login_required
@user_passes_test(is_admin)
def toggle_admin_status(request, parent_id):
    """Toggle admin status for a parent"""
    user = get_object_or_404(CustomUser, pk=parent_id, is_parent=True)
    
    # Prevent demoting yourself
    if user == request.user:
        messages.error(request, _('Vous ne pouvez pas modifier votre propre statut d\'administrateur.'))
        return redirect('admin_panel:parents_list')
    
    if request.method == 'POST':
        user.is_admin = not user.is_admin
        user.save()
        if user.is_admin:
            messages.success(request, _('%(name)s a été promu administrateur.') % {'name': f"{user.first_name} {user.last_name}"})
        else:
            messages.success(request, _('%(name)s n\'est plus administrateur.') % {'name': f"{user.first_name} {user.last_name}"})
        return redirect('admin_panel:parents_list')
    
    context = {
        'user': user,
    }
    return render(request, 'admin_panel/toggle_admin_confirm.html', context)


@login_required
@user_passes_test(is_admin)
def delete_parent_account(request, parent_id):
    """Delete a parent account"""
    user = get_object_or_404(CustomUser, pk=parent_id, is_parent=True)
    
    # Prevent deleting yourself
    if user == request.user:
        messages.error(request, _('Vous ne pouvez pas supprimer votre propre compte.'))
        return redirect('admin_panel:parents_list')
    
    if request.method == 'POST':
        user_name = f"{user.first_name} {user.last_name}"
        user.delete()
        messages.success(request, _('Le compte de %(name)s a été supprimé avec succès.') % {'name': user_name})
        return redirect('admin_panel:parents_list')
    
    context = {
        'user': user,
    }
    return render(request, 'admin_panel/delete_account_confirm.html', context)


@login_required
@user_passes_test(is_admin)
def children_list(request):
    """List all children sorted by age (ascending - youngest first)"""
    # Sort by birth_date descending (most recent = youngest = first)
    children = Child.objects.select_related('parent').order_by('-birth_date')
    
    context = {
        'children': children,
    }
    return render(request, 'admin_panel/children_list.html', context)


def welcome_page(request):
    """Display the welcome page with Markdown content"""
    welcome_page_obj = WelcomePage.get_instance()
    # Convert Markdown to HTML
    html_content = markdown.markdown(
        welcome_page_obj.content,
        extensions=['extra', 'codehilite', 'nl2br']
    )
    
    context = {
        'welcome_page': welcome_page_obj,
        'html_content': html_content,
    }
    return render(request, 'admin_panel/welcome_page.html', context)


@login_required
@user_passes_test(is_admin)
def welcome_page_edit(request):
    """Edit the welcome page content"""
    welcome_page_obj = WelcomePage.get_instance()
    
    if request.method == 'POST':
        form = WelcomePageForm(request.POST, instance=welcome_page_obj)
        if form.is_valid():
            welcome_page_obj = form.save(commit=False)
            welcome_page_obj.updated_by = request.user
            welcome_page_obj.save()
            messages.success(request, _('La page d\'accueil a été mise à jour avec succès !'))
            return redirect('admin_panel:welcome_page_edit')
    else:
        form = WelcomePageForm(instance=welcome_page_obj)
    
    # Preview the Markdown content
    html_content = markdown.markdown(
        welcome_page_obj.content,
        extensions=['extra', 'codehilite', 'nl2br']
    )
    
    context = {
        'form': form,
        'welcome_page': welcome_page_obj,
        'html_content': html_content,
    }
    return render(request, 'admin_panel/welcome_page_edit.html', context)
